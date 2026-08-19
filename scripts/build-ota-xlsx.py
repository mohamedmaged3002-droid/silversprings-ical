#!/usr/bin/env python3
"""Silver Springs OTA listing pack — modelled on Kennah OTA Listing Pack.xlsx.

Four tabs:
  Silver Springs Master — one row per unit + nightly USD by month
  Monthly Prices        — nightly USD by month
  Price Ranges          — nightly USD by continuous date range (exact, no estimation)
  Availability          — in-window occupancy per unit

USD is the primary currency: this operator quotes USD natively (the storefront
renders EGP but the rates API returns USD) and OTAs list USD, so converting to
EGP and back would only introduce error. The EGP figures the website uses live in
Supabase `unit_daily_prices`, converted at the pinned FX.

⚠️ OCCUPANCY IS COMPUTED IN-WINDOW (Brain L-119). Lodgify reports "beyond the
advance-booking window" with the same isAvailable:false it uses for "booked", so
the raw blocked count reads as ~55% forward occupancy when the truth is ~11%.
Reporting the raw number here would tell the OTA team these units are already
busy. A trailing unavailable run that reaches the horizon end AND is >=45 nights
is treated as the window edge, not a booking.

Title / status / area / guests are pulled LIVE from Supabase, not from the
scrape: the scrape is the operator's raw text, the DB is what we actually
publish. The OTA team must work from what is live or their listings drift from
the site on day one.
"""
import json, pathlib, datetime, collections, urllib.request, os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).parent.parent
PAGES = "https://mohamedmaged3002-droid.github.io/silversprings-ical"
FX = 50  # pinned, mirrors src/config.js
HORIZON_END = "2027-08-18"
WINDOW_MIN_TAIL = 45   # nights; >= this reaching the horizon end == window edge

# ---- inputs -----------------------------------------------------------------
index = json.loads((ROOT / "docs" / "index.json").read_text())
props = {p["wp"]: p for p in index["properties"]}
prices = json.loads((ROOT / "output" / "daily-prices.json").read_text())   # wp -> [{date,price(EGP),usd}]
r2 = json.loads((ROOT / "output" / "r2-photos.json").read_text())
# Per-unit Google Drive folders for the OTA team (maged@bluekeys.co, anyone-with-link).
# Sourced from the operator CDN at 4000px — NOT the 1920px R2 web copy (L-086).
_dl = ROOT / "output" / "drive-links.json"
drive = json.loads(_dl.read_text()) if _dl.exists() else {}
feeds = {int(p.stem) for p in (ROOT / "docs").glob("*.ics")}

env = {}
for line in (ROOT / ".env").read_text().splitlines():
    if "=" in line:
        k, v = line.split("=", 1); env[k.strip()] = v.strip()
req = urllib.request.Request(
    f"{env['SUPABASE_URL'].rstrip('/')}/rest/v1/units?source=eq.silversprings"
    "&select=wp_post_id,source_code,operator_unit_code,title,slug,short_description,the_property,status,"
    "beds,baths,guests,min_nights,cleaning_fee_egp,area,compound,amenities,source_url,notes,photo_urls"
    "&order=wp_post_id",
    headers={"apikey": env["SUPABASE_SERVICE_ROLE_KEY"],
             "Authorization": f"Bearer {env['SUPABASE_SERVICE_ROLE_KEY']}"})
live = json.load(urllib.request.urlopen(req, timeout=60))
print(f"live rows: {len(live)} ({sum(1 for r in live if r['status']=='published')} published)")

# ---- blocked ranges straight from the published .ics (the OTA-facing truth) ---
def ics_ranges(wp):
    f = ROOT / "docs" / f"{wp}.ics"
    if not f.exists(): return []
    starts, ends = [], []
    for ln in f.read_text().splitlines():
        if ln.startswith("DTSTART;VALUE=DATE:"): starts.append(ln.split(":")[1].strip())
        elif ln.startswith("DTEND;VALUE=DATE:"): ends.append(ln.split(":")[1].strip())
    fmt = lambda d: f"{d[:4]}-{d[4:6]}-{d[6:]}"
    return [(fmt(s), fmt(e)) for s, e in zip(starts, ends)]

def window_and_occupancy(wp):
    """-> (window_end, bookable_nights, blocked_in_window, occupancy_pct)."""
    rs = ics_ranges(wp)
    today = datetime.date.today().isoformat()
    win_end = HORIZON_END
    if rs:
        ls, le = rs[-1]
        nights = (datetime.date.fromisoformat(le) - datetime.date.fromisoformat(ls)).days
        if le >= HORIZON_END and nights >= WINDOW_MIN_TAIL:
            win_end = ls                      # window edge, not a booking (L-119)
    total = (datetime.date.fromisoformat(win_end) - datetime.date.fromisoformat(today)).days
    blocked = 0
    for s, e in rs:
        a = min(s, win_end); b = min(e, win_end)
        if b > a: blocked += (datetime.date.fromisoformat(b) - datetime.date.fromisoformat(a)).days
    occ = round(100 * blocked / total) if total > 0 else None
    return win_end, max(total, 0), blocked, occ

# ---- month buckets ----------------------------------------------------------
months, per_unit = [], {}
for wp_s, days in prices.items():
    wp = int(wp_s)
    b = collections.defaultdict(list)
    for d in days:
        b[d["date"][:7]].append(d["usd"])
    per_unit[wp] = b
    for m in b:
        if m not in months: months.append(m)
months.sort()

def label(m):
    y, mo = m.split("-")
    return f"{datetime.date(int(y), int(mo), 1):%b} '{y[2:]}"

def modal(wp, m):
    v = per_unit.get(wp, {}).get(m)
    return collections.Counter(v).most_common(1)[0][0] if v else None

# ---- styling ---------------------------------------------------------------
TITLE = Font(bold=True, size=13, color="1F4E79")
SUB = Font(italic=True, size=9, color="808080")
HDR = Font(bold=True, color="FFFFFF", size=10)
FILL = PatternFill("solid", fgColor="1F4E79")
RED = PatternFill("solid", fgColor="FFC7CE")
REDFONT = Font(color="9C0006", bold=True)

wb = Workbook(); wb.remove(wb.active)

def sheet(name, title, sub, headers, rows, widths):
    ws = wb.create_sheet(name)
    ws["A1"] = title; ws["A1"].font = TITLE
    ws["A2"] = sub;   ws["A2"].font = SUB
    ws.append([])
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(4, c); cell.font = HDR; cell.fill = FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows: ws.append(r)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{ws.max_row}"
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 28
    return ws

def why_not(r, wp, occ_total):
    reasons = []
    if wp not in feeds: reasons.append("no iCal feed generated")
    if occ_total == 0: reasons.append("ZERO availability across the whole horizon — likely switched off in Lodgify, not sold out; held pending operator confirmation")
    if r["cleaning_fee_egp"] is None: reasons.append("cleaning fee UNCONFIRMED (flat vs per-night not yet agreed) — quoting would risk under-charging")
    if not (r.get("photo_urls") or []): reasons.append("no photos")
    if r["status"] != "published": reasons.append("held as draft on bluekeys.co")
    return "; ".join(reasons)

# ---------------- Master ----------------
rows = []
for r in live:
    wp = r["wp_post_id"]
    p = props.get(wp, {})
    win_end, total, blocked, occ = window_and_occupancy(wp)
    usd_now = modal(wp, months[0]) if months else None
    photos = r.get("photo_urls") or []
    eligible = "YES" if (wp in feeds and total > 0 and photos and r["cleaning_fee_egp"] is not None
                         and r["status"] == "published") else "NO"
    rows.append([
        wp, r.get("source_code") or "", r.get("operator_unit_code") or "",
        r["title"], r.get("area") or "", r.get("compound") or "", "Apartment",
        r.get("guests"), r.get("beds"), r.get("baths"),
        usd_now, (usd_now * FX) if usd_now else None,
        r.get("min_nights") or p.get("minStay"),
        # FULL description. `short_description` is deliberately truncated to 200
        # chars for listing cards, so using it here cut every row mid-sentence.
        # The OTA team needs the whole text to build a listing from.
        (r.get("the_property") or r.get("short_description") or "").strip(),
        ", ".join(sorted(r.get("amenities") or [])),
        (drive.get(str(wp)) or {}).get("url", ""), len(photos),
        f"{PAGES}/{wp}.ics" if wp in feeds else "",
        "", "",                               # airbnb / booking — UNKNOWN, see note
        r.get("source_url") or "",
        win_end, total, blocked, occ,
        eligible, why_not(r, wp, total), r["status"],
    ] + [modal(wp, m) for m in months])

ELIG_COL = 26
rows.sort(key=lambda x: (0 if x[ELIG_COL - 1] == "YES" else 1, x[0]))

ws = sheet(
    "Silver Springs Master", "Silver Springs Residence — OTA listing pack",
    "One row per unit with the nightly USD rate for every month. 45 units, Silver Palm compound, New Cairo. "
    "photos_drive_folder links are anyone-with-link Google Drive folders (maged@bluekeys.co) holding 4000px images "
    "pulled from the operator's own CDN — NOT the smaller copies the website serves. "
    "occupancy_pct is computed INSIDE each unit's advance-booking window — the raw blocked count would read ~5x higher "
    "because Lodgify reports 'beyond the booking window' identically to 'booked' (Brain L-119). "
    "RED ROWS AT THE BOTTOM ARE NOT READY TO LIST — why_not_ready says why. "
    "airbnb / booking_com are DELIBERATELY BLANK: we do not yet know whether the operator already lists these units "
    "on those channels under their own account. Confirm before listing anything, or you risk double-booking.",
    ["wp_post_id", "code", "unit_code", "title", "area", "compound", "property_type",
     "guests", "bedrooms", "bathrooms", "nightly_usd", "nightly_egp", "min_stay",
     "description", "amenities", "photos_drive_folder", "photo_count",
     "ical_url", "airbnb", "booking_com", "source_url",
     "bookable_until", "bookable_nights", "blocked_in_window", "occupancy_pct",
     "ota_eligible", "why_not_ready", "bluekeys_status"] + [label(m) + " USD" for m in months],
    rows,
    [11, 8, 14, 44, 12, 13, 13, 8, 10, 10, 11, 11, 9, 84, 46, 30, 12, 50, 30, 30, 46,
     13, 14, 16, 12, 13, 74, 15] + [11] * len(months))

for rr in range(5, ws.max_row + 1):
    ws.cell(rr, 14).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(rr, 27).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[rr].height = 76
    if ws.cell(rr, ELIG_COL).value != "YES":
        for c in range(1, len(ws[4]) + 1):
            ws.cell(rr, c).fill = RED
            ws.cell(rr, c).font = REDFONT

# ---------------- Monthly Prices ----------------
rows = []
for r in live:
    wp = r["wp_post_id"]
    if not per_unit.get(wp): continue
    rows.append([wp, r.get("source_code") or "", r["title"], r.get("area") or "", r.get("beds")]
                + [modal(wp, m) for m in months])
sheet("Monthly Prices", "Nightly price by month — USD",
      "One row per unit. Each month shows the most common nightly USD rate in that month. Blank = no rate published.",
      ["wp", "code", "title", "area", "beds"] + [label(m) for m in months],
      rows, [10, 8, 42, 13, 7] + [11] * len(months))

# ---------------- Price Ranges ----------------
rows = []
for r in live:
    wp = r["wp_post_id"]
    days = sorted(prices.get(str(wp), []), key=lambda d: d["date"])
    if not days: continue
    run = None
    def flush(run):
        if run:
            n = (datetime.date.fromisoformat(run[1]) - datetime.date.fromisoformat(run[0])).days + 1
            rows.append([wp, r.get("source_code") or "", r["title"], r.get("beds"),
                         run[0], run[1], n, run[2], round(run[2] * FX)])
    for d in days:
        if run and d["usd"] == run[2] and \
           (datetime.date.fromisoformat(d["date"]) - datetime.date.fromisoformat(run[1])).days == 1:
            run[1] = d["date"]
        else:
            flush(run); run = [d["date"], d["date"], d["usd"]]
    flush(run)
sheet("Price Ranges", "Nightly price by date range — USD",
      "Each row is a continuous run of dates at one flat nightly USD rate. Exact, no estimation. EGP shown at the pinned FX of 50.",
      ["wp", "code", "title", "beds", "from", "to", "nights", "nightly_usd", "nightly_egp"],
      rows, [10, 8, 42, 7, 12, 12, 8, 13, 13])

# ---------------- Availability ----------------
rows = []
for r in live:
    wp = r["wp_post_id"]
    win_end, total, blocked, occ = window_and_occupancy(wp)
    raw_blocked = sum(
        (datetime.date.fromisoformat(e) - datetime.date.fromisoformat(s)).days
        for s, e in ics_ranges(wp))
    rows.append([wp, r.get("source_code") or "", r["title"], r.get("beds"),
                 win_end, total, blocked, occ, raw_blocked,
                 "window-limited" if win_end < HORIZON_END else "full horizon"])
rows.sort(key=lambda x: -(x[7] or 0))
sheet("Availability", "Availability — in-window only",
      "occupancy_pct = blocked_in_window / bookable_nights. raw_blocked_all_horizon is shown ONLY to make the gap "
      "visible: it counts nights beyond the operator's advance-booking window, which are not bookable and are NOT "
      "occupancy. Never quote raw_blocked as occupancy (Brain L-119).",
      ["wp", "code", "title", "beds", "bookable_until", "bookable_nights",
       "blocked_in_window", "occupancy_pct", "raw_blocked_all_horizon", "window"],
      rows, [10, 8, 42, 7, 14, 15, 17, 13, 22, 16])

out = pathlib.Path("/Users/MAGED/inv/Silver Springs OTA Listing Pack.xlsx")
wb.save(out)
print(f"saved {out}")
for s in wb.sheetnames:
    print(f"  {s}: {wb[s].max_row - 4} rows")
elig = sum(1 for r in ws.iter_rows(min_row=5, values_only=True) if r[ELIG_COL - 1] == "YES")
print(f"ota_eligible YES: {elig}/{len(live)}")
