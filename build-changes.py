#!/usr/bin/env python3
"""Build the "changes only" xlsx from out/changed-units.json (written by pricewatch.js).

Attached to the change email by send-alert.js, which then sends an EMPTY body — the
sheet IS the message. Mirrors almaza-ical/build-changes.py and kennah-ical's
equivalent, with two deliberate differences:

  1. NO MARKUP. Almaza's sheet multiplies by 1.10 to match the markup baked into its
     OTA sheet. Silver Springs is 0% (service_fee_percent = 0, no cleaning fee), so
     the numbers here are the operator's own — the same figures their storefront and
     ours both show. Adding a markup column here would invent one.

  2. ROSTER CHANGES GO IN THE SHEET, on their own tab. Almaza leaves added/removed
     units in the email body only, which means once the attachment rule sends an
     empty body, a roster-only day arrives as a blank email. Keeping them here makes
     every alert attachment-only and self-contained.

Prices: the operator quotes USD, so USD is authoritative and EGP is shown alongside
at the run's FX (their own live rate). No file is written — and any stale one is
removed — when nothing changed.
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "out/changed-units.json"
OUT = "out/silversprings-changes.xlsx"

HDR_FILL = PatternFill("solid", fgColor="FF1F3B57")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
BASE = Font(name="Arial", size=10)
THIN = Side(style="thin", color="FFD9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style(ws, ncols, money_cols=(), delta_col=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 26
    for ri in range(2, ws.max_row + 1):
        for ci in range(1, ncols + 1):
            cell = ws.cell(ri, ci)
            cell.font = BASE
            cell.border = BORDER
            if ci in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0;-#,##0"
        if delta_col:
            chg = ws.cell(ri, delta_col)
            if isinstance(chg.value, (int, float)):
                chg.font = Font(name="Arial", size=10, bold=True,
                                color="FFB00020" if chg.value < 0 else "FF1B7A3D")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + ncols)}{ws.max_row}"


def main():
    if not os.path.exists(SRC):
        print("build-changes: no changed-units.json — nothing to build.")
        return
    data = json.load(open(SRC))
    price_units = data.get("units") or []
    added = data.get("addedUnits") or []
    removed = data.get("removedUnits") or []

    if not price_units and not added and not removed:
        print("build-changes: no changes — not writing a sheet.")
        if os.path.exists(OUT):
            os.remove(OUT)
        return

    fx = data.get("fx") or 0
    date_str = data.get("dateStr", "")
    wb = Workbook()
    wb.remove(wb.active)

    # ---- Price Changes -------------------------------------------------------
    if price_units:
        cols = [("Unit ID", 11), ("Code", 9), ("Unit", 42),
                ("From", 13), ("To", 13),
                ("Old USD", 11), ("New USD", 11), ("Change USD", 12),
                ("Old EGP", 13), ("New EGP", 13)]
        ws = wb.create_sheet("Price Changes")
        ws.append([c[0] for c in cols])
        for i, (_, w) in enumerate(cols, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        rows = []
        for u in price_units:
            for r in u.get("ranges", []):
                # `oldEgp`/`newEgp` are the diff's field names but carry USD here —
                # the baseline stores the quoted currency (see pricewatch.js).
                old_usd, new_usd = r.get("oldEgp"), r.get("newEgp")
                delta = (new_usd - old_usd) if isinstance(old_usd, (int, float)) and isinstance(new_usd, (int, float)) else None
                rows.append([u.get("wp"), u.get("code"), u.get("title"),
                             r.get("from"), r.get("to"),
                             old_usd, new_usd, delta,
                             round(old_usd * fx) if isinstance(old_usd, (int, float)) and fx else None,
                             round(new_usd * fx) if isinstance(new_usd, (int, float)) and fx else None])
        rows.sort(key=lambda x: (str(x[0]), str(x[3])))
        for row in rows:
            ws.append(row)
        style(ws, len(cols), money_cols=(6, 7, 8, 9, 10), delta_col=8)
        n_ranges = len(rows)
    else:
        n_ranges = 0

    # ---- Roster Changes ------------------------------------------------------
    if added or removed:
        cols = [("Change", 12), ("Slug on their site", 62), ("What to do", 66)]
        ws = wb.create_sheet("Roster Changes")
        ws.append([c[0] for c in cols])
        for i, (_, w) in enumerate(cols, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        for s in added:
            ws.append(["NEW", s, "Not on BlueKeys yet — onboard it (content.js, build-roster, build-insert-sql)."])
        for s in removed:
            ws.append(["REMOVED", s, "Gone from their site — we may still be selling it. Check, then draft or delist."])
        style(ws, len(cols))
        for ri in range(2, ws.max_row + 1):
            cell = ws.cell(ri, 1)
            cell.font = Font(name="Arial", size=10, bold=True,
                             color="FF1B7A3D" if cell.value == "NEW" else "FFB00020")
            ws.cell(ri, 3).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT)
    print(f"build-changes: wrote {OUT} — {n_ranges} changed date-range(s) across "
          f"{len(price_units)} unit(s), +{len(added)} / -{len(removed)} roster, FX {fx}, {date_str}.")


if __name__ == "__main__":
    main()
